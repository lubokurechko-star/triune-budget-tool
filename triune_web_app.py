import streamlit as st
import os, io, json, hashlib
from datetime import datetime
import pandas as pd
 
# Configure matplotlib BEFORE importing pyplot
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
 
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
 
# ── Brand Colors ──────────────────────────────────────────────
NAVY="#1F3864"; TEAL="#2E75B6"; LIGHT="#D6E4F0"; WHITE="#FFFFFF"
GREEN="#70AD47"; RED="#C00000"; DARK="#152848"; PURPLE="#9B59B6"
GOLD="#FFF2CC"
 
OX_NAVY="1F3864"; OX_TEAL="2E75B6"; OX_LIGHT="D6E4F0"
OX_WHITE="FFFFFF"; OX_GOLD="FFF2CC"
 
# ── Page Configuration ────────────────────────────────────────
st.set_page_config(
    page_title="Triune Budget Analysis Tool",
    page_icon="🎭",
    layout="wide",
    initial_sidebar_state="expanded"
)
 
# ── Custom CSS ─────────────────────────────────────────────────
st.markdown("""
    <style>
    .main {background-color: #F0F5FB;}
    .stButton>button {
        background-color: #2E75B6; color: white; font-weight: bold;
        border-radius: 5px; padding: 10px 20px; border: none;
    }
    .stButton>button:hover {background-color: #1F3864;}
    h1 {color: #1F3864; font-family: Georgia, serif;}
    h2 {color: #2E75B6; font-family: Georgia, serif;}
    </style>
    """, unsafe_allow_html=True)
 
 
# ═══════════════════════════════════════════════════
#  PASSWORD PROTECTION
# ═══════════════════════════════════════════════════
 
def check_password():
    """Returns True if user entered correct password."""
    
    if "password_correct" not in st.session_state:
        st.session_state["password_correct"] = False
    
    if st.session_state["password_correct"]:
        return True
    
    st.markdown("""
        <div style='background-color: #1F3864; padding: 30px; border-radius: 10px; margin-bottom: 20px; text-align: center;'>
            <h1 style='color: white; margin: 0;'> Triune Entertainment</h1>
            <h2 style='color: #D6E4F0; margin: 10px 0;'>Budget Analysis Tool</h2>
        </div>
    """, unsafe_allow_html=True)
    
    st.markdown("### Secure Login")
    st.markdown("Please enter your password to access the Budget Analysis Tool.")
    
    password = st.text_input("Password", type="password", key="password_input")
    
    col1, col2, col3 = st.columns([1, 1, 1])
    with col2:
        if st.button("Login", use_container_width=True):
            if password == "triune2024":
                st.session_state["password_correct"] = True
                st.rerun()
            else:
                st.error(" Incorrect password. Please try again.")
    
    st.markdown("---")
    st.info("Contact Team 4 for login information.")
    
    return False
 
 
# ═══════════════════════════════════════════════════
#  DATA EXTRACTION
# ═══════════════════════════════════════════════════
 
def extract_budget_data(uploaded_file):
    """Extract Budget (LEFT col 7) and Actual (RIGHT col 13) data."""
    try:
        df = pd.read_excel(uploaded_file, header=None)
    except Exception as e:
        return None, f"Cannot read file: {e}"
    
    # Extract show name from row 2, column 2
    show_name = "Unknown Show"
    show_date = ""
    
    # Check row 2 for show name
    if df.shape[0] > 2:
        for col in range(df.shape[1]):
            val = df.iloc[2, col]
            if pd.notna(val) and isinstance(val, str) and len(val) > 5:
                show_name = val.strip()
                if " - " in show_name:
                    parts = show_name.split(" - ")
                    show_name = parts[0].strip()
                    show_date = parts[1].strip() if len(parts) > 1 else ""
                if "Director:" in show_name:
                    show_name = show_name.split("Director:")[0].strip()
                break
    
    # Extract revenue: LEFT (col 7) = BUDGET, RIGHT (col 13) = ACTUAL
    budget_revenue = actual_revenue = 0
    for idx in range(len(df)):
        row_text = str(df.iloc[idx, 1]) if pd.notna(df.iloc[idx, 1]) else ""
        if "Total 4300 Revenues" in row_text:
            budget_revenue = pd.to_numeric(df.iloc[idx, 7], errors='coerce') or 0
            actual_revenue = pd.to_numeric(df.iloc[idx, 13], errors='coerce') or 0
            break
    
    # Extract expenses: LEFT (col 7) = BUDGET, RIGHT (col 13) = ACTUAL
    budget_expenses = actual_expenses = 0
    for idx in range(len(df)):
        row_text = str(df.iloc[idx, 1]) if pd.notna(df.iloc[idx, 1]) else ""
        if "Total 5000 Direct Production Costs" in row_text:
            budget_expenses = pd.to_numeric(df.iloc[idx, 7], errors='coerce') or 0
            actual_expenses = pd.to_numeric(df.iloc[idx, 13], errors='coerce') or 0
            break
    
    # Calculate metrics
    budget_net = budget_revenue - budget_expenses
    actual_net = actual_revenue - actual_expenses
    budget_margin = (budget_net / budget_revenue * 100) if budget_revenue > 0 else 0
    actual_margin = (actual_net / actual_revenue * 100) if actual_revenue > 0 else 0
    revenue_variance = actual_revenue - budget_revenue
    expense_variance = actual_expenses - budget_expenses
    net_variance = actual_net - budget_net
    
    # Extract detailed expense categories
    expense_categories = []
    in_expense_section = False
    
    for idx in range(len(df)):
        row_text = str(df.iloc[idx, 1]) if pd.notna(df.iloc[idx, 1]) else ""
        
        # Start tracking when we hit 5000 section
        if "5000 Direct Production Costs" in row_text:
            in_expense_section = True
            continue
        
        # Stop when we hit the total
        if "Total 5000" in row_text:
            break
        
        # Extract category details
        if in_expense_section:
            cat_name = str(df.iloc[idx, 2]) if pd.notna(df.iloc[idx, 2]) else ""
            if cat_name and cat_name.strip() and cat_name != 'nan':
                cat_budget = pd.to_numeric(df.iloc[idx, 7], errors='coerce') or 0
                cat_actual = pd.to_numeric(df.iloc[idx, 13], errors='coerce') or 0
                cat_variance = cat_actual - cat_budget
                cat_variance_pct = (cat_variance / cat_budget * 100) if cat_budget > 0 else 0
                
                # Clean up category name (remove code number)
                if ' ' in cat_name:
                    parts = cat_name.split(' ', 1)
                    if len(parts) > 1:
                        cat_name_clean = parts[1].strip()
                    else:
                        cat_name_clean = cat_name
                else:
                    cat_name_clean = cat_name
                
                expense_categories.append({
                    'name': cat_name_clean,
                    'budget': cat_budget,
                    'actual': cat_actual,
                    'variance': cat_variance,
                    'variance_pct': cat_variance_pct
                })
    
    return {
        'filename': uploaded_file.name,
        'show_name': show_name,
        'show_date': show_date,
        'budget_revenue': budget_revenue,
        'actual_revenue': actual_revenue,
        'budget_expenses': budget_expenses,
        'actual_expenses': actual_expenses,
        'budget_net': budget_net,
        'actual_net': actual_net,
        'budget_margin': budget_margin,
        'actual_margin': actual_margin,
        'revenue_variance': revenue_variance,
        'expense_variance': expense_variance,
        'net_variance': net_variance,
        'revenue_variance_pct': (revenue_variance / budget_revenue * 100) if budget_revenue > 0 else 0,
        'expense_variance_pct': (expense_variance / budget_expenses * 100) if budget_expenses > 0 else 0,
        'expense_categories': expense_categories
    }, None
 
 
# ═══════════════════════════════════════════════════
#  CHART FUNCTIONS (6 total)
# ═══════════════════════════════════════════════════
 
def create_chart_1_budget_vs_actual(data):
    """Chart 1: Budget vs Actual - 3 comparisons"""
    fig, (ax1, ax2, ax3) = plt.subplots(1, 3, figsize=(18, 6))
    fig.patch.set_facecolor('#F0F5FB')
    
    x = [0, 1]
    
    # Revenue
    ax1.set_facecolor('#FFFFFF')
    bars1 = ax1.bar(x, [data['budget_revenue'], data['actual_revenue']],
                    color=[TEAL, PURPLE], alpha=0.9, width=0.6, linewidth=2, edgecolor='white')
    ax1.set_title('Revenue: Budget vs Actual', fontsize=14, fontweight='bold', color=NAVY, pad=15)
    ax1.set_xticks(x)
    ax1.set_xticklabels(['Budget', 'Actual'])
    ax1.yaxis.set_major_formatter(mticker.FuncFormatter(lambda v, _: f'${v:,.0f}'))
    ax1.grid(axis='y', alpha=0.3, linestyle='--')
    for bar in bars1:
        h = bar.get_height()
        ax1.annotate(f'${h:,.0f}', xy=(bar.get_x()+bar.get_width()/2, h),
                    xytext=(0,5), textcoords='offset points', ha='center', fontsize=10, fontweight='bold')
    
    # Expenses
    ax2.set_facecolor('#FFFFFF')
    bars2 = ax2.bar(x, [data['budget_expenses'], data['actual_expenses']],
                    color=[NAVY, RED], alpha=0.9, width=0.6, linewidth=2, edgecolor='white')
    ax2.set_title('Expenses: Budget vs Actual', fontsize=14, fontweight='bold', color=NAVY, pad=15)
    ax2.set_xticks(x)
    ax2.set_xticklabels(['Budget', 'Actual'])
    ax2.yaxis.set_major_formatter(mticker.FuncFormatter(lambda v, _: f'${v:,.0f}'))
    ax2.grid(axis='y', alpha=0.3, linestyle='--')
    for bar in bars2:
        h = bar.get_height()
        ax2.annotate(f'${h:,.0f}', xy=(bar.get_x()+bar.get_width()/2, h),
                    xytext=(0,5), textcoords='offset points', ha='center', fontsize=10, fontweight='bold')
    
    # Net Income
    ax3.set_facecolor('#FFFFFF')
    colors = [TEAL if v>=0 else RED for v in [data['budget_net'], data['actual_net']]]
    bars3 = ax3.bar(x, [data['budget_net'], data['actual_net']],
                    color=colors, alpha=0.9, width=0.6, linewidth=2, edgecolor='white')
    ax3.set_title('Net Income: Budget vs Actual', fontsize=14, fontweight='bold', color=NAVY, pad=15)
    ax3.set_xticks(x)
    ax3.set_xticklabels(['Budget', 'Actual'])
    ax3.yaxis.set_major_formatter(mticker.FuncFormatter(lambda v, _: f'${v:,.0f}'))
    ax3.axhline(y=0, color='black', linewidth=1)
    ax3.grid(axis='y', alpha=0.3, linestyle='--')
    for bar in bars3:
        h = bar.get_height()
        ax3.annotate(f'${h:,.0f}', xy=(bar.get_x()+bar.get_width()/2, h),
                    xytext=(0,5 if h>=0 else -15), textcoords='offset points', ha='center', fontsize=10, fontweight='bold')
    
    plt.tight_layout()
    return fig
 
def create_chart_2_variance(data):
    """Chart 2: Variance Analysis"""
    fig, ax = plt.subplots(figsize=(14, 9))
    fig.patch.set_facecolor('#F0F5FB')
    ax.set_facecolor('#FFFFFF')
    
    categories = ['Revenue', 'Expenses', 'Net Income']
    variances = [data['revenue_variance'], data['expense_variance'], data['net_variance']]
    colors = [GREEN if v>=0 else RED for v in variances]
    
    bars = ax.bar(categories, variances, color=colors, alpha=0.9, width=0.6, linewidth=2, edgecolor='white')
    ax.set_title('Variance Analysis (Actual - Budget)', fontsize=18, fontweight='bold', color=NAVY, pad=25)
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda v, _: f'${v:,.0f}'))
    ax.axhline(y=0, color='black', linewidth=2)
    ax.grid(True, alpha=0.3, linestyle='--', linewidth=0.8)
    ax.set_ylabel('Variance Amount ($)', fontsize=14, fontweight='bold', labelpad=10)
    ax.set_xlabel('Category', fontsize=14, fontweight='bold', labelpad=10)
    
    for bar, val in zip(bars, variances):
        h = bar.get_height()
        pct = (val / data['budget_revenue'] * 100) if data['budget_revenue'] > 0 else 0
        # Position label further from bar to prevent overlap
        y_offset = 15 if h >= 0 else -45
        ax.annotate(f'${abs(h):,.0f}\n({abs(pct):.1f}%)',
                   xy=(bar.get_x()+bar.get_width()/2, h),
                   xytext=(0, y_offset), textcoords='offset points',
                   ha='center', fontsize=13, fontweight='bold')
    
    # Add more y-axis padding
    y_max = max(abs(min(variances)), abs(max(variances)))
    ax.set_ylim(-y_max * 1.3, y_max * 1.3)
    
    plt.tight_layout(pad=2.0)
    return fig
 
def create_chart_3_pie(data):
    """Chart 3: Pie Charts"""
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(16, 7))
    fig.patch.set_facecolor('#F0F5FB')
    
    # Budget Pie
    sizes1 = [data['budget_net'], data['budget_expenses']]
    labels1 = [f'Net ${sizes1[0]:,.0f}', f'Expenses ${sizes1[1]:,.0f}']
    wedges1, texts1, autotexts1 = ax1.pie(sizes1, labels=labels1, autopct='%1.1f%%',
            colors=[TEAL, NAVY], explode=(0.1,0), shadow=True, startangle=90)
    ax1.set_title(f'Budget\nTotal: ${data["budget_revenue"]:,.0f}', fontsize=14, fontweight='bold', pad=15)
    
    # Actual Pie
    sizes2 = [data['actual_net'], data['actual_expenses']]
    labels2 = [f'Net ${sizes2[0]:,.0f}', f'Expenses ${sizes2[1]:,.0f}']
    wedges2, texts2, autotexts2 = ax2.pie(sizes2, labels=labels2, autopct='%1.1f%%',
            colors=[PURPLE, RED], explode=(0.1,0), shadow=True, startangle=90)
    ax2.set_title(f'Actual\nTotal: ${data["actual_revenue"]:,.0f}', fontsize=14, fontweight='bold', pad=15)
    
    plt.tight_layout()
    return fig
 
def create_chart_4_scatter(data):
    """Chart 4: Scatter Plot"""
    fig, ax = plt.subplots(figsize=(11, 8))
    fig.patch.set_facecolor('#F0F5FB')
    ax.set_facecolor('#FFFFFF')
    
    categories = ['Revenue', 'Expenses', 'Net']
    budget = [data['budget_revenue'], data['budget_expenses'], data['budget_net']]
    actual = [data['actual_revenue'], data['actual_expenses'], data['actual_net']]
    colors = [TEAL, NAVY, PURPLE]
    
    for cat, b, a, c in zip(categories, budget, actual, colors):
        ax.scatter(b, a, s=300, c=c, alpha=0.7, edgecolors='white', linewidth=2, label=cat)
        ax.annotate(cat, (b,a), xytext=(10,10), textcoords='offset points', fontsize=9)
    
    max_val = max(budget + actual) * 1.1
    min_val = min(budget + actual) * 0.9
    ax.plot([min_val,max_val], [min_val,max_val], 'k--', alpha=0.5, linewidth=2, label='Perfect')
    
    ax.set_xlabel('Budget ($)', fontsize=12, fontweight='bold')
    ax.set_ylabel('Actual ($)', fontsize=12, fontweight='bold')
    ax.set_title('Budget Accuracy', fontsize=15, fontweight='bold', color=NAVY, pad=15)
    ax.xaxis.set_major_formatter(mticker.FuncFormatter(lambda v, _: f'${v:,.0f}'))
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda v, _: f'${v:,.0f}'))
    ax.legend()
    ax.grid(True, alpha=0.3)
    
    plt.tight_layout()
    return fig
 
def create_chart_5_line(data):
    """Chart 5: Line Graph"""
    fig, ax = plt.subplots(figsize=(15, 9))
    fig.patch.set_facecolor('#F0F5FB')
    ax.set_facecolor('#FFFFFF')
    
    categories = ['Revenue', 'Expenses', 'Net Income']
    x_pos = [0,1,2]
    budget = [data['budget_revenue'], data['budget_expenses'], data['budget_net']]
    actual = [data['actual_revenue'], data['actual_expenses'], data['actual_net']]
    
    ax.plot(x_pos, budget, marker='o', markersize=14, linewidth=4, color=TEAL,
            label='Budget', markeredgecolor='white', markeredgewidth=3)
    ax.plot(x_pos, actual, marker='s', markersize=14, linewidth=4, color=PURPLE,
            label='Actual', markeredgecolor='white', markeredgewidth=3, linestyle='--')
    
    for i, (b,a) in enumerate(zip(budget, actual)):
        # Budget labels - position higher
        ax.annotate(f'${b:,.0f}', (i,b), xytext=(0,18), textcoords='offset points',
                   ha='center', fontsize=11, fontweight='bold', color=TEAL,
                   bbox=dict(boxstyle='round,pad=0.4', facecolor='white', 
                            edgecolor=TEAL, linewidth=2))
        # Actual labels - position lower
        ax.annotate(f'${a:,.0f}', (i,a), xytext=(0,-25), textcoords='offset points',
                   ha='center', fontsize=11, fontweight='bold', color=PURPLE,
                   bbox=dict(boxstyle='round,pad=0.4', facecolor='white', 
                            edgecolor=PURPLE, linewidth=2))
    
    ax.set_xticks(x_pos)
    ax.set_xticklabels(categories, fontsize=13, fontweight='bold')
    ax.set_ylabel('Amount ($)', fontsize=14, fontweight='bold', labelpad=10)
    ax.set_title('Trend Analysis', fontsize=18, fontweight='bold', color=NAVY, pad=20)
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda v, _: f'${v:,.0f}'))
    ax.legend(fontsize=13, frameon=True, shadow=True)
    ax.grid(True, alpha=0.3, axis='y', linewidth=0.8)
    
    # Add more y-axis padding for labels
    y_min = min(min(budget), min(actual))
    y_max = max(max(budget), max(actual))
    y_range = y_max - y_min
    ax.set_ylim(y_min - y_range*0.15, y_max + y_range*0.15)
    
    plt.tight_layout(pad=2.0)
    return fig
 
def create_chart_6_bar(data):
    """Chart 6: Comprehensive Bar Graph"""
    fig, ax = plt.subplots(figsize=(15, 8))
    fig.patch.set_facecolor('#F0F5FB')
    ax.set_facecolor('#FFFFFF')
    
    categories = ['Revenue', 'Expenses', 'Net Income']
    budget = [data['budget_revenue'], data['budget_expenses'], data['budget_net']]
    actual = [data['actual_revenue'], data['actual_expenses'], data['actual_net']]
    
    x = range(len(categories))
    width = 0.35
    
    bars1 = ax.bar([i-width/2 for i in x], budget, width, label='Budget', color=TEAL, alpha=0.9,
                   edgecolor='white', linewidth=2)
    bars2 = ax.bar([i+width/2 for i in x], actual, width, label='Actual', color=PURPLE, alpha=0.9,
                   edgecolor='white', linewidth=2)
    
    for bars in [bars1, bars2]:
        for bar in bars:
            h = bar.get_height()
            ax.annotate(f'${h:,.0f}', xy=(bar.get_x()+bar.get_width()/2, h),
                       xytext=(0, 5 if h>=0 else -15), textcoords='offset points',
                       ha='center', fontsize=10, fontweight='bold')
    
    # Variance labels
    for i, (b,a) in enumerate(zip(budget, actual)):
        if b != 0:
            var_pct = ((a-b)/b)*100
            color = GREEN if var_pct>=0 else RED
            symbol = "▲" if var_pct>=0 else "▼"
            ax.annotate(f'{symbol} {abs(var_pct):.1f}%', xy=(i, max(b,a)),
                       xytext=(0,25), textcoords='offset points', ha='center',
                       fontsize=12, fontweight='bold', color=color)
    
    ax.set_xlabel('Category', fontsize=13, fontweight='bold')
    ax.set_ylabel('Amount ($)', fontsize=13, fontweight='bold')
    ax.set_title('Complete Financial Comparison', fontsize=16, fontweight='bold', color=NAVY, pad=20)
    ax.set_xticks(x)
    ax.set_xticklabels(categories, fontsize=12, fontweight='bold')
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda v, _: f'${v:,.0f}'))
    ax.legend(fontsize=12)
    ax.grid(True, alpha=0.3, axis='y')
    ax.axhline(y=0, color='black', linewidth=1)
    
    plt.tight_layout()
    return fig
 
 
def create_chart_7_category_breakdown(data):
    """Chart 7: Detailed Category Breakdown"""
    if not data.get('expense_categories'):
        # Return empty figure if no categories
        fig, ax = plt.subplots(figsize=(14, 8))
        ax.text(0.5, 0.5, 'No category data available', ha='center', va='center')
        return fig
    
    categories = data['expense_categories']
    
    # Sort by variance (biggest overruns first)
    sorted_cats = sorted(categories, key=lambda x: x['variance'], reverse=True)
    
    # Take top 10 categories
    top_cats = sorted_cats[:10]
    
    # Increase figure size for better spacing
    fig, ax = plt.subplots(figsize=(20, 12))
    fig.patch.set_facecolor('#F0F5FB')
    ax.set_facecolor('#FFFFFF')
    
    names = [cat['name'][:50] for cat in top_cats]  # Allow longer names
    budgets = [cat['budget'] for cat in top_cats]
    actuals = [cat['actual'] for cat in top_cats]
    variances = [cat['variance'] for cat in top_cats]
    
    y_pos = range(len(names))
    height = 0.4  # Slightly thinner bars for more spacing
    
    # Create horizontal bars with more spacing
    bars1 = ax.barh([i-height/2 for i in y_pos], budgets, height, 
                    label='Budget', color=TEAL, alpha=0.9, edgecolor='white', linewidth=2)
    bars2 = ax.barh([i+height/2 for i in y_pos], actuals, height,
                    label='Actual', color=PURPLE, alpha=0.9, edgecolor='white', linewidth=2)
    
    # Add value labels with better positioning
    for bars in [bars1, bars2]:
        for bar in bars:
            width = bar.get_width()
            if width > 0:
                ax.annotate(f'${width:,.0f}',
                           xy=(width, bar.get_y() + bar.get_height()/2),
                           xytext=(8, 0), textcoords='offset points',
                           va='center', fontsize=11, fontweight='bold')
    
    # Add variance indicators with better spacing
    max_val = max(max(budgets), max(actuals))
    for i, (cat, var) in enumerate(zip(top_cats, variances)):
        if cat['budget'] > 0:
            var_pct = cat['variance_pct']
            color = RED if var > 0 else GREEN
            symbol = "▲" if var > 0 else "▼"
            
            # Position variance label further to the right
            x_pos = max(cat['budget'], cat['actual']) + (max_val * 0.15)
            
            ax.annotate(f'{symbol} ${abs(var):,.0f} ({abs(var_pct):.1f}%)',
                       xy=(x_pos, i),
                       xytext=(0, 0), textcoords='offset points',
                       va='center', fontsize=12, fontweight='bold', color=color,
                       bbox=dict(boxstyle='round,pad=0.5', facecolor='white',
                                edgecolor=color, linewidth=2.5))
    
    ax.set_yticks(y_pos)
    ax.set_yticklabels(names, fontsize=11)
    ax.set_xlabel('Amount ($)', fontsize=14, fontweight='bold', labelpad=10)
    ax.set_title('Top 10 Expense Categories: Budget vs Actual\n(Red ▲ = Over Budget | Green ▼ = Under Budget)', 
                 fontsize=18, fontweight='bold', color=NAVY, pad=25)
    ax.xaxis.set_major_formatter(mticker.FuncFormatter(lambda v, _: f'${v:,.0f}'))
    ax.legend(fontsize=13, loc='lower right', frameon=True, shadow=True)
    ax.grid(True, alpha=0.3, axis='x', linewidth=0.8)
    
    # Add more padding to prevent label cutoff
    ax.margins(x=0.25)
    
    # Invert y-axis so biggest variance is on top
    ax.invert_yaxis()
    
    plt.tight_layout(pad=2.0)
    return fig
 
 
# ═══════════════════════════════════════════════════
#  EXCEL REPORT
# ═══════════════════════════════════════════════════
 
def generate_excel_report(data, charts_dict):
    """Generate Excel report with embedded charts."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    
    ws['A1'] = f"Show: {data['show_name']}"
    ws['A1'].font = Font(bold=True, size=16, color=OX_NAVY)
    ws['A2'] = f"Date: {data['show_date']}"
    ws['A3'] = f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
    
    ws['A5'] = "Category"
    ws['B5'] = "Budget"
    ws['C5'] = "Actual"
    ws['D5'] = "Variance"
    ws['E5'] = "Variance %"
    
    for col in ['A','B','C','D','E']:
        ws[f'{col}5'].fill = PatternFill('solid', fgColor=OX_TEAL)
        ws[f'{col}5'].font = Font(bold=True, color=OX_WHITE)
        ws[f'{col}5'].alignment = Alignment(horizontal='center')
    
    ws['A6'] = "Revenue"
    ws['B6'] = data['budget_revenue']
    ws['C6'] = data['actual_revenue']
    ws['D6'] = data['revenue_variance']
    ws['E6'] = data['revenue_variance_pct']/100
    
    ws['A7'] = "Expenses"
    ws['B7'] = data['budget_expenses']
    ws['C7'] = data['actual_expenses']
    ws['D7'] = data['expense_variance']
    ws['E7'] = data['expense_variance_pct']/100
    
    ws['A8'] = "Net Income"
    ws['B8'] = data['budget_net']
    ws['C8'] = data['actual_net']
    ws['D8'] = data['net_variance']
    ws['E8'] = (data['net_variance']/data['budget_net'])/100 if data['budget_net']!=0 else 0
    
    for row in range(6,9):
        for col in ['B','C','D']:
            ws[f'{col}{row}'].number_format = '$#,##0.00'
        ws[f'E{row}'].number_format = '0.00%'
        if row%2==0:
            for col in ['A','B','C','D','E']:
                ws[f'{col}{row}'].fill = PatternFill('solid', fgColor=OX_LIGHT)
    
    for col, width in [('A',15), ('B',18), ('C',18), ('D',18), ('E',15)]:
        ws.column_dimensions[col].width = width
    
    # Add Top 5 Overruns and Savings sections
    if data.get('expense_categories'):
        cats = data['expense_categories']
        overruns = sorted([c for c in cats if c['variance'] > 0], 
                         key=lambda x: x['variance'], reverse=True)[:5]
        savings = sorted([c for c in cats if c['variance'] < 0], 
                        key=lambda x: x['variance'])[:5]
        
        # Top 5 Overruns section
        current_row = 11
        ws[f'A{current_row}'] = "TOP 5 BUDGET OVERRUNS"
        ws[f'A{current_row}'].font = Font(bold=True, size=14, color=OX_NAVY)
        ws.merge_cells(f'A{current_row}:E{current_row}')
        
        current_row += 1
        ws[f'A{current_row}'] = "Category"
        ws[f'B{current_row}'] = "Budget"
        ws[f'C{current_row}'] = "Actual"
        ws[f'D{current_row}'] = "Variance"
        ws[f'E{current_row}'] = "Variance %"
        
        for col in ['A','B','C','D','E']:
            ws[f'{col}{current_row}'].fill = PatternFill('solid', fgColor='C00000')  # Red header
            ws[f'{col}{current_row}'].font = Font(bold=True, color=OX_WHITE)
            ws[f'{col}{current_row}'].alignment = Alignment(horizontal='center')
        
        if overruns:
            for i, cat in enumerate(overruns, 1):
                current_row += 1
                ws[f'A{current_row}'] = f"{i}. {cat['name']}"
                ws[f'B{current_row}'] = cat['budget']
                ws[f'C{current_row}'] = cat['actual']
                ws[f'D{current_row}'] = cat['variance']
                ws[f'E{current_row}'] = cat['variance_pct']/100
                
                for col in ['B','C','D']:
                    ws[f'{col}{current_row}'].number_format = '$#,##0.00'
                ws[f'E{current_row}'].number_format = '0.0%'
                
                # Alternate row colors
                if i%2==0:
                    for col in ['A','B','C','D','E']:
                        ws[f'{col}{current_row}'].fill = PatternFill('solid', fgColor=OX_LIGHT)
        else:
            current_row += 1
            ws[f'A{current_row}'] = "✓ No categories over budget!"
            ws.merge_cells(f'A{current_row}:E{current_row}')
            ws[f'A{current_row}'].font = Font(color='70AD47', bold=True)
        
        # Top 5 Savings section
        current_row += 2
        ws[f'A{current_row}'] = "TOP 5 BUDGET SAVINGS"
        ws[f'A{current_row}'].font = Font(bold=True, size=14, color=OX_NAVY)
        ws.merge_cells(f'A{current_row}:E{current_row}')
        
        current_row += 1
        ws[f'A{current_row}'] = "Category"
        ws[f'B{current_row}'] = "Budget"
        ws[f'C{current_row}'] = "Actual"
        ws[f'D{current_row}'] = "Variance"
        ws[f'E{current_row}'] = "Variance %"
        
        for col in ['A','B','C','D','E']:
            ws[f'{col}{current_row}'].fill = PatternFill('solid', fgColor='70AD47')  # Green header
            ws[f'{col}{current_row}'].font = Font(bold=True, color=OX_WHITE)
            ws[f'{col}{current_row}'].alignment = Alignment(horizontal='center')
        
        if savings:
            for i, cat in enumerate(savings, 1):
                current_row += 1
                ws[f'A{current_row}'] = f"{i}. {cat['name']}"
                ws[f'B{current_row}'] = cat['budget']
                ws[f'C{current_row}'] = cat['actual']
                ws[f'D{current_row}'] = cat['variance']
                ws[f'E{current_row}'] = cat['variance_pct']/100
                
                for col in ['B','C','D']:
                    ws[f'{col}{current_row}'].number_format = '$#,##0.00'
                ws[f'E{current_row}'].number_format = '0.0%'
                
                # Alternate row colors
                if i%2==0:
                    for col in ['A','B','C','D','E']:
                        ws[f'{col}{current_row}'].fill = PatternFill('solid', fgColor=OX_LIGHT)
        else:
            current_row += 1
            ws[f'A{current_row}'] = "No categories under budget"
            ws.merge_cells(f'A{current_row}:E{current_row}')
    
    # Adjust column A width to fit longer category names
    ws.column_dimensions['A'].width = 50
    
    # Add each chart to separate sheets
    chart_names = ['Budget_vs_Actual', 'Variance_Analysis', 'Pie_Charts', 
                   'Scatter_Plot', 'Line_Graph', 'Bar_Graph', 'Category_Breakdown']
    
    for chart_name, fig in zip(chart_names, charts_dict.values()):
        # Create new sheet
        ws_chart = wb.create_sheet(title=chart_name)
        
        # Save figure to bytes
        img_buffer = io.BytesIO()
        fig.savefig(img_buffer, format='png', dpi=150, bbox_inches='tight')
        img_buffer.seek(0)
        
        # Add image to sheet
        img = XLImage(img_buffer)
        ws_chart.add_image(img, 'A1')
    
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
 
 
# ═══════════════════════════════════════════════════
#  MAIN APP
# ═══════════════════════════════════════════════════
 
def main():
    if not check_password():
        return
    
    st.markdown("""
        <div style='background-color: #1F3864; padding: 20px; border-radius: 10px; margin-bottom: 20px;'>
            <h1 style='color: white; text-align: center; margin: 0;'>Triune Entertainment</h1>
            <h2 style='color: #D6E4F0; text-align: center; margin: 5px 0 15px 0;'>Budget Analysis & Visualization Tool</h2>
            <div style='text-align: center;'>
                <a href='https://ads.google.com/aw/campaigns/new/performancemax?campaignId=281498546159968&ocid=8030665422' 
                   target='_blank' 
                   style='background-color: #2E75B6; color: white; padding: 10px 20px; margin: 0 10px; 
                          text-decoration: none; border-radius: 5px; font-weight: bold; display: inline-block;'>
                     Google Ads
                </a>
                <a href='https://airtable.com/appDuTE72UfHIFOJT/shrmRPv812yX1gDlZ/tblDl6p4SNVzHLQmZ/viw0wLiFVeOZYvJ6b' 
                   target='_blank' 
                   style='background-color: #70AD47; color: white; padding: 10px 20px; margin: 0 10px; 
                          text-decoration: none; border-radius: 5px; font-weight: bold; display: inline-block;'>
                     Talent CRM
                </a>
            </div>
        </div>
    """, unsafe_allow_html=True)
    
    with st.sidebar:
        st.markdown("###  Upload Budget Worksheets")
        uploaded_files = st.file_uploader("Choose Excel files", type=['xlsx','xls'],
                                         accept_multiple_files=True)
        st.markdown("---")
        st.markdown("### 📊 Features")
        st.markdown(" 7 detailed charts\n, Variance analysis\n, Multi-file support\n, and Password protected")
        
        if st.button(" Logout"):
            st.session_state["password_correct"] = False
            st.rerun()
    
    if not uploaded_files:
        st.info(" Upload budget Excel sheets to get started")
    else:
        st.markdown("###  Analysis Results")
        
        for idx, uploaded_file in enumerate(uploaded_files):
            st.markdown("---")
            with st.expander(f" {uploaded_file.name}", expanded=True):
                data, error = extract_budget_data(uploaded_file)
                
                if error:
                    st.error(f" Error: {error}")
                    continue
                
                st.markdown(f"## {data['show_name']} ({data['show_date']})")
                
                col1, col2, col3, col4 = st.columns(4)
                
                # Pre-format all values
                budget_rev_str = f"${data['budget_revenue']:,.2f}"
                actual_rev_str = f"${data['actual_revenue']:,.2f}"
                rev_delta_str = f"${data['revenue_variance']:,.2f}"
                
                budget_exp_str = f"${data['budget_expenses']:,.2f}"
                actual_exp_str = f"${data['actual_expenses']:,.2f}"
                exp_delta_str = f"${data['expense_variance']:,.2f}"
                
                budget_net_str = f"${data['budget_net']:,.2f}"
                actual_net_str = f"${data['actual_net']:,.2f}"
                net_delta_str = f"${data['net_variance']:,.2f}"
                
                budget_margin_str = f"{data['budget_margin']:.1f}%"
                actual_margin_str = f"{data['actual_margin']:.1f}%"
                margin_diff = data['actual_margin'] - data['budget_margin']
                margin_delta_str = f"{margin_diff:.1f}%"
                
                with col1:
                    st.metric("Budget Revenue", budget_rev_str)
                    delta_color = "normal" if data['revenue_variance'] >= 0 else "inverse"
                    st.metric("Actual Revenue", actual_rev_str,
                             delta=rev_delta_str, delta_color=delta_color)
                
                with col2:
                    st.metric("Budget Expenses", budget_exp_str)
                    delta_color = "inverse" if data['expense_variance'] >= 0 else "normal"
                    st.metric("Actual Expenses", actual_exp_str,
                             delta=exp_delta_str, delta_color=delta_color)
                
                with col3:
                    st.metric("Budget Net", budget_net_str)
                    delta_color = "normal" if data['net_variance'] >= 0 else "inverse"
                    st.metric("Actual Net", actual_net_str,
                             delta=net_delta_str, delta_color=delta_color)
                
                with col4:
                    st.metric("Budget Margin", budget_margin_str)
                    delta_color = "normal" if margin_diff >= 0 else "inverse"
                    st.metric("Actual Margin", actual_margin_str,
                             delta=margin_delta_str, delta_color=delta_color)
                
                # Category-level insights
                if data.get('expense_categories'):
                    st.markdown("###  Category Performance Insights")
                    
                    cats = data['expense_categories']
                    
                    # Sort by variance
                    overruns = sorted([c for c in cats if c['variance'] > 0], 
                                    key=lambda x: x['variance'], reverse=True)[:5]
                    savings = sorted([c for c in cats if c['variance'] < 0], 
                                   key=lambda x: x['variance'])[:5]
                    
                    col_over, col_save = st.columns(2)
                    
                    with col_over:
                        st.markdown("####  Top 5 Budget Overruns")
                        if overruns:
                            for i, cat in enumerate(overruns, 1):
                                st.markdown(f"""
                                **{i}. {cat['name']}**  
                                Budget: ${cat['budget']:,.2f} | Actual: ${cat['actual']:,.2f}  
                                <span style='color: #C00000; font-weight: bold;'>
                                ▲ ${cat['variance']:,.2f} (+{cat['variance_pct']:.1f}%)
                                </span>
                                """, unsafe_allow_html=True)
                        else:
                            st.success(" No categories over budget!")
                    
                    with col_save:
                        st.markdown("####  Top 5 Budget Savings")
                        if savings:
                            for i, cat in enumerate(savings, 1):
                                st.markdown(f"""
                                **{i}. {cat['name']}**  
                                Budget: ${cat['budget']:,.2f} | Actual: ${cat['actual']:,.2f}  
                                <span style='color: #70AD47; font-weight: bold;'>
                                ▼ ${abs(cat['variance']):,.2f} (-{abs(cat['variance_pct']):.1f}%)
                                </span>
                                """, unsafe_allow_html=True)
                        else:
                            st.info("No categories under budget")
                
                st.markdown("###  All 7 Detailed Charts")
                
                tab1,tab2,tab3,tab4,tab5,tab6,tab7 = st.tabs([
                    "Budget vs Actual", "Variance", "Pie Charts", 
                    "Scatter Plot", "Line Graph", "Bar Graph", "Category Breakdown"
                ])
                
                # Create all charts and store them
                fig1 = create_chart_1_budget_vs_actual(data)
                fig2 = create_chart_2_variance(data)
                fig3 = create_chart_3_pie(data)
                fig4 = create_chart_4_scatter(data)
                fig5 = create_chart_5_line(data)
                fig6 = create_chart_6_bar(data)
                fig7 = create_chart_7_category_breakdown(data)
                
                charts_dict = {
                    'budget_vs_actual': fig1,
                    'variance': fig2,
                    'pie': fig3,
                    'scatter': fig4,
                    'line': fig5,
                    'bar': fig6,
                    'category_breakdown': fig7
                }
                
                with tab1:
                    st.pyplot(fig1)
                with tab2:
                    st.pyplot(fig2)
                with tab3:
                    st.pyplot(fig3)
                with tab4:
                    st.pyplot(fig4)
                with tab5:
                    st.pyplot(fig5)
                with tab6:
                    st.pyplot(fig6)
                with tab7:
                    st.pyplot(fig7)
                
                # Close all figures after display
                for fig in charts_dict.values():
                    plt.close(fig)
                
                st.markdown("###  Download Report")
                
                # Custom filename input with unique key
                default_filename = f"{data['show_name']}_{data['show_date']}_analysis"
                custom_filename = st.text_input(
                    "Report filename (without .xlsx extension):",
                    value=default_filename,
                    help="Enter a custom name for your report or use the default",
                    key=f"filename_input_{idx}_{uploaded_file.name}"
                )
                
                # Generate Excel with charts
                excel_data = generate_excel_report(data, charts_dict)
                st.download_button(
                    label=f" Download {data['show_name']} Excel Report (with all 7 charts)",
                    data=excel_data,
                    file_name=f"{custom_filename}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key=f"download_btn_{idx}_{uploaded_file.name}"
                )
 
if __name__ == "__main__":
    main()
